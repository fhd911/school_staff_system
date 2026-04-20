import re
import zipfile
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import FieldDoesNotExist, PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from accounts.decorators import get_current_supervisor, supervisor_login_required
from accounts.models import Supervisor

from .filters import apply_principal_filters, apply_vice_filters, build_filter_choices
from .forms import (
    CorrectionDecisionForm,
    CorrectionRequestForm,
    PrincipalRecordForm,
    SupervisorAdminUpdateForm,
    SupervisorImportForm,
    VicePrincipalRecordForm,
)
from .models import (
    AccountResetRequest,
    CorrectionRequest,
    DataEntryWindow,
    PrincipalRecord,
    VicePrincipalRecord,
)
from .services import build_admin_overview_context


def _clean_filename(value):
    value = (value or "").strip()
    value = re.sub(r'[\\/*?:"<>|]+', "-", value)
    value = re.sub(r"\s+", " ", value)
    return value or "file"


def _model_has_field(model, field_name):
    try:
        model._meta.get_field(field_name)
        return True
    except FieldDoesNotExist:
        return False


def _count_active(queryset):
    if _model_has_field(queryset.model, "is_active"):
        return queryset.filter(is_active=True).count()
    return 0


def _count_attachments(queryset):
    if not _model_has_field(queryset.model, "performance_file"):
        return 0
    return queryset.exclude(performance_file="").exclude(performance_file__isnull=True).count()


def _collect_school_count(principal_qs, vice_qs):
    school_names = set(principal_qs.values_list("school_name", flat=True))
    school_names.update(vice_qs.values_list("school_name", flat=True))
    return len([name for name in school_names if name])


def _build_current_filters(request):
    return {
        "q": request.GET.get("q", "").strip(),
        "sector": request.GET.get("sector", "").strip(),
        "stage": request.GET.get("stage", "").strip(),
        "school_gender": request.GET.get("school_gender", "").strip(),
        "supervisor": request.GET.get("supervisor", "").strip(),
        "is_active": request.GET.get("is_active", "").strip(),
        "has_attachment": request.GET.get("has_attachment", "").strip(),
    }


def _build_admin_filter_context(request):
    principal_base_qs = PrincipalRecord.objects.select_related("supervisor").all()
    vice_base_qs = VicePrincipalRecord.objects.select_related("supervisor").all()

    return {
        "filter_choices": build_filter_choices(principal_base_qs, vice_base_qs),
        "stage_options": PrincipalRecord.STAGE_CHOICES,
        "school_gender_options": PrincipalRecord.SCHOOL_GENDER_CHOICES,
        "current_filters": _build_current_filters(request),
    }


def _build_correction_initial_from_record(record):
    return {
        "requested_full_name": getattr(record, "full_name", ""),
        "requested_national_id": getattr(record, "national_id", ""),
        "requested_mobile": getattr(record, "mobile", ""),
        "requested_school_name": getattr(record, "school_name", ""),
        "requested_sector": getattr(record, "sector", ""),
        "requested_stage": getattr(record, "stage", ""),
        "requested_school_gender": getattr(record, "school_gender", ""),
        "requested_role": getattr(record, "role", ""),
        "requested_notes": getattr(record, "notes", ""),
    }


def _has_open_correction_request_for_record(record, target_type):
    qs = CorrectionRequest.objects.filter(
        target_type=target_type,
        status=CorrectionRequest.STATUS_PENDING,
    )
    if target_type == CorrectionRequest.TARGET_PRINCIPAL:
        qs = qs.filter(principal_record=record)
    else:
        qs = qs.filter(vice_record=record)
    return qs.exists()


def _build_supervisor_bound_form(request, form_class, supervisor):
    instance = form_class._meta.model(supervisor=supervisor)
    return form_class(
        request.POST or None,
        request.FILES or None,
        instance=instance,
    )


def _build_supervisor_update_form(request, form_class, instance):
    return form_class(
        request.POST or None,
        request.FILES or None,
        instance=instance,
    )


def _get_active_entry_window():
    now = timezone.now()
    return (
        DataEntryWindow.objects.filter(
            is_active=True,
            starts_at__lte=now,
            ends_at__gte=now,
        )
        .order_by("-starts_at")
        .first()
    )


def _get_latest_entry_window():
    return DataEntryWindow.objects.order_by("-starts_at").first()


def _build_entry_window_context(supervisor=None):
    now = timezone.now()
    window = _get_active_entry_window()
    latest_window = window or _get_latest_entry_window()

    status_label = "لا توجد فترة مسجلة"
    can_add = False
    can_edit = False
    can_delete = False

    if latest_window:
        if not latest_window.is_active:
            status_label = "غير مفعلة"
        elif now < latest_window.starts_at:
            status_label = "لم تبدأ"
        elif latest_window.starts_at <= now <= latest_window.ends_at:
            status_label = "مفتوحة"
        else:
            status_label = "منتهية"

        can_add = status_label == "مفتوحة" and getattr(latest_window, "allow_add", False)
        can_edit = status_label == "مفتوحة" and getattr(latest_window, "allow_edit", False)
        can_delete = status_label == "مفتوحة" and getattr(latest_window, "allow_delete", False)

    if supervisor:
        can_add = can_add and getattr(supervisor, "can_add_records", True)
        can_edit = can_edit and getattr(supervisor, "can_edit_records", False)
        can_delete = can_delete and getattr(supervisor, "can_delete_records", False)

    return {
        "entry_window": latest_window,
        "entry_window_status_label": status_label,
        "entry_window_can_add": can_add,
        "entry_window_can_edit": can_edit,
        "entry_window_can_delete": can_delete,
        "entry_window_remaining_seconds": getattr(latest_window, "remaining_seconds", 0) if latest_window else 0,
        "entry_window_remaining_days": getattr(latest_window, "remaining_days", 0) if latest_window else 0,
        "entry_window_remaining_hours": getattr(latest_window, "remaining_hours", 0) if latest_window else 0,
    }


def _ensure_supervisor_can_add(supervisor):
    if not getattr(supervisor, "can_add_records", True):
        raise PermissionDenied("ليست لديك صلاحية إضافة السجلات.")

    window = _get_active_entry_window()
    if not window:
        raise PermissionDenied("فترة تسجيل البيانات مغلقة حاليًا.")

    if not getattr(window, "allow_add", False):
        raise PermissionDenied("الإضافة غير متاحة خلال الفترة الحالية.")

    return window


def _ensure_supervisor_can_edit(supervisor):
    if not getattr(supervisor, "can_edit_records", False):
        raise PermissionDenied("ليست لديك صلاحية تعديل السجلات.")

    window = _get_active_entry_window()
    if not window:
        raise PermissionDenied("فترة تعديل البيانات مغلقة حاليًا.")

    if not getattr(window, "allow_edit", False):
        raise PermissionDenied("التعديل غير متاح خلال الفترة الحالية.")

    return window


def _ensure_supervisor_can_delete(supervisor):
    if not getattr(supervisor, "can_delete_records", False):
        raise PermissionDenied("ليست لديك صلاحية حذف السجلات.")

    window = _get_active_entry_window()
    if not window:
        raise PermissionDenied("فترة حذف البيانات مغلقة حاليًا.")

    if not getattr(window, "allow_delete", False):
        raise PermissionDenied("الحذف غير متاح خلال الفترة الحالية.")

    return window


def _create_correction_request(request, record, target_type, title):
    supervisor = get_current_supervisor(request)

    if _has_open_correction_request_for_record(record, target_type):
        messages.warning(request, "يوجد بالفعل طلب تصحيح مفتوح لهذا السجل.")
        return redirect("staffdata:records_list")

    form = CorrectionRequestForm(
        request.POST or None,
        request.FILES or None,
        initial=_build_correction_initial_from_record(record),
        target_type=target_type,
    )

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.supervisor = supervisor
        obj.target_type = target_type
        obj.status = CorrectionRequest.STATUS_PENDING

        if target_type == CorrectionRequest.TARGET_PRINCIPAL:
            obj.principal_record = record
        else:
            obj.vice_record = record

        obj.save()
        messages.success(request, "تم رفع طلب التصحيح بنجاح وإحالته للإدارة للمراجعة.")
        return redirect("staffdata:my_correction_requests")

    return render(
        request,
        "staffdata/correction_request_form.html",
        {
            "form": form,
            "page_title": title,
            "record": record,
        },
    )


def _apply_correction_to_record(correction_request):
    if correction_request.target_type == CorrectionRequest.TARGET_PRINCIPAL:
        record = correction_request.principal_record
    else:
        record = correction_request.vice_record

    if record is None:
        raise ValueError("السجل المرتبط بطلب التصحيح غير موجود.")

    record.full_name = correction_request.requested_full_name
    record.national_id = correction_request.requested_national_id
    record.mobile = getattr(correction_request, "requested_mobile", "")
    record.school_name = correction_request.requested_school_name
    record.sector = correction_request.requested_sector
    record.stage = correction_request.requested_stage
    record.school_gender = correction_request.requested_school_gender
    record.role = correction_request.requested_role
    record.notes = correction_request.requested_notes

    requested_file = getattr(correction_request, "requested_performance_file", None)
    if requested_file and hasattr(record, "performance_file"):
        record.performance_file = requested_file

    record.save()
    return record


def _safe_supervisor_value(supervisor, attr_name, default="—"):
    value = getattr(supervisor, attr_name, None)
    if value in (None, ""):
        return default
    return value


def _fallback_aware_datetime():
    return timezone.make_aware(
        datetime(2000, 1, 1, 0, 0, 0),
        timezone.get_current_timezone(),
    )


def _merge_latest_activity(*values):
    valid_values = [value for value in values if value]
    if not valid_values:
        return None
    return max(valid_values)


def _build_supervisor_status(
    total_count,
    missing_attachments,
    pending_corrections_count,
    pending_reset_requests_count=0,
):
    if (
        pending_corrections_count > 0
        or missing_attachments > 0
        or pending_reset_requests_count > 0
    ):
        return "يحتاج متابعة", "warn"
    if total_count > 0:
        return "نشط", "good"
    return "دون مدخلات", "idle"


def _reset_supervisor_account(supervisor):
    update_fields = []

    if _model_has_field(Supervisor, "password"):
        supervisor.password = ""
        update_fields.append("password")

    if _model_has_field(Supervisor, "is_activated"):
        supervisor.is_activated = False
        update_fields.append("is_activated")

    if _model_has_field(Supervisor, "password_set_at"):
        supervisor.password_set_at = None
        update_fields.append("password_set_at")

    if _model_has_field(Supervisor, "last_login_at"):
        supervisor.last_login_at = None
        update_fields.append("last_login_at")

    if update_fields:
        supervisor.save(update_fields=update_fields)
    else:
        supervisor.save()


def _auto_fit_xlsx_columns(ws, rows):
    widths = {}
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            text = str(value or "")
            widths[col_idx] = max(widths.get(col_idx, 0), len(text))

    for col_idx, width in widths.items():
        adjusted = min(max(width + 4, 12), 35)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _build_xlsx_response(filename, sheet_title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Sheet1")[:31]

    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4D3A")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    body_font = Font(color="1F1F1F", size=10)
    thin_side = Side(style="thin", color="D9E3DC")
    body_border = Border(bottom=thin_side)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = body_border

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = body_border

    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24
    _auto_fit_xlsx_columns(ws, [headers, *rows])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _normalize_digits(value):
    return "".join(filter(str.isdigit, str(value or "").strip()))


def _normalize_mobile_for_import(value):
    digits = _normalize_digits(value)

    if digits.startswith("966") and len(digits) == 12:
        digits = "0" + digits[3:]
    elif digits.startswith("5") and len(digits) == 9:
        digits = "0" + digits

    return digits


def _normalize_bool_from_excel(value):
    if value in (None, ""):
        return None

    if value in (True, 1):
        return True

    if value in (False, 0):
        return False

    value_str = str(value).strip().lower()

    if value_str in {"1", "true", "yes", "y", "نعم", "نشط", "active"}:
        return True

    if value_str in {"0", "false", "no", "n", "لا", "غير نشط", "inactive"}:
        return False

    return None


def _normalize_import_header(header):
    return str(header or "").strip().lower().replace(" ", "_")


def _resolve_import_headers(header_row):
    raw_headers = {}
    for idx, header in enumerate(header_row):
        normalized = _normalize_import_header(header)
        if normalized:
            raw_headers[normalized] = idx

    aliases = {
        "full_name": {
            "full_name", "fullname", "name",
            "الاسم", "الاسم_الرباعي", "اسم_المشرف", "اسم_المشرفة",
        },
        "national_id": {
            "national_id", "id", "civil_id", "nationalid",
            "السجل_المدني", "رقم_الهوية", "الهوية", "السجل",
        },
        "mobile": {
            "mobile", "phone", "phone_number", "mobile_number",
            "الجوال", "رقم_الجوال", "رقم_التواصل",
        },
        "email": {
            "email", "البريد_الإلكتروني", "البريد_الالكتروني", "الايميل", "الإيميل",
        },
        "is_active": {
            "is_active", "active", "status",
            "نشط", "الحالة",
        },
        "sector": {
            "sector", "القطاع",
        },
        "can_add_records": {
            "can_add_records", "allow_add", "add_permission",
            "صلاحية_الإضافة", "سماح_الإضافة", "يستطيع_الإضافة",
        },
        "can_edit_records": {
            "can_edit_records", "allow_edit", "edit_permission",
            "صلاحية_التعديل", "سماح_التعديل", "يستطيع_التعديل",
        },
        "can_delete_records": {
            "can_delete_records", "allow_delete", "delete_permission",
            "صلاحية_الحذف", "سماح_الحذف", "يستطيع_الحذف",
        },
    }

    resolved = {}
    for canonical_name, possible_names in aliases.items():
        for possible_name in possible_names:
            if possible_name in raw_headers:
                resolved[canonical_name] = raw_headers[possible_name]
                break

    return resolved


@staff_member_required(login_url="/admin/login/")
@require_POST
@transaction.atomic
def admin_reset_supervisor_account_view(request, supervisor_id):
    supervisor = get_object_or_404(Supervisor, pk=supervisor_id)

    _reset_supervisor_account(supervisor)

    pending_request = (
        AccountResetRequest.objects.filter(
            supervisor=supervisor,
            status=AccountResetRequest.STATUS_PENDING,
        )
        .order_by("-created_at")
        .first()
    )

    if pending_request:
        pending_request.status = AccountResetRequest.STATUS_PROCESSED
        pending_request.processed_at = timezone.now()
        pending_request.processed_by = request.user
        pending_request.save(update_fields=["status", "processed_at", "processed_by"])

    messages.success(
        request,
        f"تمت إعادة تهيئة حساب المشرف {supervisor.full_name} بنجاح."
    )

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")
    if next_url:
        return redirect(next_url)

    return redirect("staffdata:admin_supervisor_detail", supervisor_id=supervisor.id)


@staff_member_required(login_url="/admin/login/")
@transaction.atomic
def admin_import_supervisors_view(request):
    form = SupervisorImportForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        excel_file = form.cleaned_data["file"]
        update_existing = form.cleaned_data["update_existing"]
        reset_existing_accounts = form.cleaned_data["reset_existing_accounts"]

        workbook = load_workbook(excel_file, data_only=True)
        worksheet = workbook.active

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "الملف المرفوع فارغ.")
            return redirect("staffdata:admin_import_supervisors")

        headers = _resolve_import_headers(rows[0])

        required_headers = ["full_name", "national_id", "mobile"]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            messages.error(
                request,
                f"الملف لا يحتوي على الأعمدة المطلوبة: {', '.join(missing_headers)}"
            )
            return redirect("staffdata:admin_import_supervisors")

        created_count = 0
        updated_count = 0
        skipped_count = 0
        reset_count = 0
        error_rows = []

        for excel_row_number, row in enumerate(rows[1:], start=2):
            if not row or all(value in (None, "") for value in row):
                continue

            full_name = str(row[headers["full_name"]] or "").strip()
            national_id = _normalize_digits(row[headers["national_id"]] or "")
            mobile = _normalize_mobile_for_import(row[headers["mobile"]] or "")

            email = ""
            if "email" in headers:
                email = str(row[headers["email"]] or "").strip()

            sector = ""
            if "sector" in headers:
                sector = str(row[headers["sector"]] or "").strip()

            raw_is_active = None
            if "is_active" in headers:
                raw_is_active = _normalize_bool_from_excel(row[headers["is_active"]])

            raw_can_add_records = None
            if "can_add_records" in headers:
                raw_can_add_records = _normalize_bool_from_excel(row[headers["can_add_records"]])

            raw_can_edit_records = None
            if "can_edit_records" in headers:
                raw_can_edit_records = _normalize_bool_from_excel(row[headers["can_edit_records"]])

            raw_can_delete_records = None
            if "can_delete_records" in headers:
                raw_can_delete_records = _normalize_bool_from_excel(row[headers["can_delete_records"]])

            is_active = True if raw_is_active is None else raw_is_active
            can_add_records = True if raw_can_add_records is None else raw_can_add_records
            can_edit_records = False if raw_can_edit_records is None else raw_can_edit_records
            can_delete_records = False if raw_can_delete_records is None else raw_can_delete_records

            if not full_name or not national_id or not mobile:
                skipped_count += 1
                error_rows.append(f"الصف {excel_row_number}: بيانات ناقصة.")
                continue

            if len(national_id) != 10:
                skipped_count += 1
                error_rows.append(f"الصف {excel_row_number}: السجل المدني يجب أن يكون 10 أرقام.")
                continue

            if not (mobile.startswith("05") and len(mobile) == 10):
                skipped_count += 1
                error_rows.append(f"الصف {excel_row_number}: رقم الجوال غير صحيح.")
                continue

            supervisor = Supervisor.objects.filter(national_id=national_id).first()

            if supervisor is None:
                create_kwargs = {}

                if _model_has_field(Supervisor, "full_name"):
                    create_kwargs["full_name"] = full_name
                if _model_has_field(Supervisor, "national_id"):
                    create_kwargs["national_id"] = national_id
                if _model_has_field(Supervisor, "mobile"):
                    create_kwargs["mobile"] = mobile
                if _model_has_field(Supervisor, "email"):
                    create_kwargs["email"] = email
                if _model_has_field(Supervisor, "sector"):
                    create_kwargs["sector"] = sector
                if _model_has_field(Supervisor, "is_active"):
                    create_kwargs["is_active"] = is_active
                if _model_has_field(Supervisor, "can_add_records"):
                    create_kwargs["can_add_records"] = can_add_records
                if _model_has_field(Supervisor, "can_edit_records"):
                    create_kwargs["can_edit_records"] = can_edit_records
                if _model_has_field(Supervisor, "can_delete_records"):
                    create_kwargs["can_delete_records"] = can_delete_records
                if _model_has_field(Supervisor, "password"):
                    create_kwargs["password"] = ""
                if _model_has_field(Supervisor, "is_activated"):
                    create_kwargs["is_activated"] = False
                if _model_has_field(Supervisor, "password_set_at"):
                    create_kwargs["password_set_at"] = None
                if _model_has_field(Supervisor, "last_login_at"):
                    create_kwargs["last_login_at"] = None

                Supervisor.objects.create(**create_kwargs)
                created_count += 1
                continue

            if not update_existing:
                skipped_count += 1
                continue

            update_fields = []

            if _model_has_field(Supervisor, "full_name") and supervisor.full_name != full_name:
                supervisor.full_name = full_name
                update_fields.append("full_name")

            if _model_has_field(Supervisor, "mobile") and getattr(supervisor, "mobile", "") != mobile:
                supervisor.mobile = mobile
                update_fields.append("mobile")

            if _model_has_field(Supervisor, "email") and getattr(supervisor, "email", "") != email:
                supervisor.email = email
                update_fields.append("email")

            if _model_has_field(Supervisor, "sector") and getattr(supervisor, "sector", "") != sector:
                supervisor.sector = sector
                update_fields.append("sector")

            if (
                _model_has_field(Supervisor, "is_active")
                and raw_is_active is not None
                and getattr(supervisor, "is_active", True) != raw_is_active
            ):
                supervisor.is_active = raw_is_active
                update_fields.append("is_active")

            if (
                _model_has_field(Supervisor, "can_add_records")
                and raw_can_add_records is not None
                and getattr(supervisor, "can_add_records", True) != raw_can_add_records
            ):
                supervisor.can_add_records = raw_can_add_records
                update_fields.append("can_add_records")

            if (
                _model_has_field(Supervisor, "can_edit_records")
                and raw_can_edit_records is not None
                and getattr(supervisor, "can_edit_records", False) != raw_can_edit_records
            ):
                supervisor.can_edit_records = raw_can_edit_records
                update_fields.append("can_edit_records")

            if (
                _model_has_field(Supervisor, "can_delete_records")
                and raw_can_delete_records is not None
                and getattr(supervisor, "can_delete_records", False) != raw_can_delete_records
            ):
                supervisor.can_delete_records = raw_can_delete_records
                update_fields.append("can_delete_records")

            if update_fields:
                supervisor.save(update_fields=update_fields)
                updated_count += 1
            else:
                skipped_count += 1

            if reset_existing_accounts:
                _reset_supervisor_account(supervisor)
                reset_count += 1

        if created_count:
            messages.success(request, f"تم إنشاء {created_count} مشرف/مشرفة.")
        if updated_count:
            messages.success(request, f"تم تحديث {updated_count} مشرف/مشرفة.")
        if reset_count:
            messages.warning(request, f"تمت إعادة تهيئة {reset_count} حساب/حسابات.")
        if skipped_count:
            messages.info(request, f"تم تخطي {skipped_count} صف/صفوف.")

        if error_rows:
            messages.warning(request, " | ".join(error_rows[:10]))

        return redirect("staffdata:admin_import_supervisors")

    return render(
        request,
        "staffdata/admin_import_supervisors.html",
        {
            "page_title": "استيراد المشرفين",
            "form": form,
        },
    )


@staff_member_required(login_url="/admin/login/")
def admin_download_supervisors_template_view(request):
    headers = [
        "full_name",
        "national_id",
        "mobile",
        "email",
        "is_active",
        "sector",
        "can_add_records",
        "can_edit_records",
        "can_delete_records",
    ]
    rows = [
        ["محمد أحمد علي", "1234567890", "0501234567", "m@example.com", 1, "أبها", 1, 0, 0],
        ["سارة خالد حسن", "2345678901", "0507654321", "s@example.com", 1, "خميس مشيط", 1, 1, 0],
    ]

    return _build_xlsx_response(
        filename="supervisors_import_template.xlsx",
        sheet_title="نموذج المشرفين",
        headers=headers,
        rows=rows,
    )


@supervisor_login_required
def dashboard_view(request):
    supervisor = get_current_supervisor(request)

    principal_records = (
        PrincipalRecord.objects.filter(supervisor=supervisor, is_active=True)
        .select_related("supervisor")
        .order_by("-created_at")
    )
    vice_records = (
        VicePrincipalRecord.objects.filter(supervisor=supervisor, is_active=True)
        .select_related("supervisor")
        .order_by("-created_at")
    )

    principal_count = principal_records.count()
    vice_count = vice_records.count()
    total_count = principal_count + vice_count
    attachments_count = _count_attachments(principal_records)

    recent_items = []

    for item in principal_records[:5]:
        recent_items.append(
            {
                "type": "مدير/مديرة",
                "name": item.full_name,
                "school": item.school_name,
                "sector": item.sector,
                "role": item.role,
                "created_at": item.created_at,
            }
        )

    for item in vice_records[:5]:
        recent_items.append(
            {
                "type": "وكيل/وكيلة",
                "name": item.full_name,
                "school": item.school_name,
                "sector": item.sector,
                "role": item.role,
                "created_at": item.created_at,
            }
        )

    recent_items.sort(key=lambda x: x["created_at"], reverse=True)
    recent_items = recent_items[:8]

    context = {
        "supervisor": supervisor,
        "principal_count": principal_count,
        "vice_count": vice_count,
        "total_count": total_count,
        "attachments_count": attachments_count,
        "recent_items": recent_items,
        **_build_entry_window_context(supervisor),
    }

    return render(request, "staffdata/dashboard.html", context)


@supervisor_login_required
def principal_create_view(request):
    supervisor = get_current_supervisor(request)

    try:
        _ensure_supervisor_can_add(supervisor)
    except PermissionDenied as exc:
        messages.error(request, str(exc))
        return redirect("staffdata:dashboard")

    form = _build_supervisor_bound_form(request, PrincipalRecordForm, supervisor)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.supervisor = supervisor
        obj.save()
        messages.success(request, "تم حفظ سجل المدير/المديرة بنجاح.")
        return redirect("staffdata:records_list")

    return render(
        request,
        "staffdata/principal_form.html",
        {
            "form": form,
            "page_title": "إضافة مدير/مديرة",
            "submit_label": "حفظ البيانات",
            **_build_entry_window_context(supervisor),
        },
    )


@supervisor_login_required
def vice_create_view(request):
    supervisor = get_current_supervisor(request)

    try:
        _ensure_supervisor_can_add(supervisor)
    except PermissionDenied as exc:
        messages.error(request, str(exc))
        return redirect("staffdata:dashboard")

    form = _build_supervisor_bound_form(request, VicePrincipalRecordForm, supervisor)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.supervisor = supervisor
        obj.save()
        messages.success(request, "تم حفظ سجل الوكيل/الوكيلة بنجاح.")
        return redirect("staffdata:records_list")

    return render(
        request,
        "staffdata/vice_form.html",
        {
            "form": form,
            "page_title": "إضافة وكيل/وكيلة",
            "submit_label": "حفظ البيانات",
            **_build_entry_window_context(supervisor),
        },
    )


@supervisor_login_required
def records_list_view(request):
    supervisor = get_current_supervisor(request)

    principal_records = (
        PrincipalRecord.objects.filter(supervisor=supervisor, is_active=True)
        .select_related("supervisor")
        .order_by("-created_at")
    )
    vice_records = (
        VicePrincipalRecord.objects.filter(supervisor=supervisor, is_active=True)
        .select_related("supervisor")
        .order_by("-created_at")
    )

    principal_count = principal_records.count()
    vice_count = vice_records.count()
    attachments_count = _count_attachments(principal_records)
    open_correction_requests_count = CorrectionRequest.objects.filter(
        supervisor=supervisor,
        status=CorrectionRequest.STATUS_PENDING,
    ).count()

    return render(
        request,
        "staffdata/records_list.html",
        {
            "principal_records": principal_records,
            "vice_records": vice_records,
            "principal_count": principal_count,
            "vice_count": vice_count,
            "total_count": principal_count + vice_count,
            "attachments_count": attachments_count,
            "open_correction_requests_count": open_correction_requests_count,
            **_build_entry_window_context(supervisor),
        },
    )


@supervisor_login_required
def principal_update_view(request, pk):
    supervisor = get_current_supervisor(request)
    record = get_object_or_404(PrincipalRecord, pk=pk, supervisor=supervisor, is_active=True)

    try:
        _ensure_supervisor_can_edit(supervisor)
    except PermissionDenied as exc:
        messages.error(request, str(exc))
        return redirect("staffdata:records_list")

    form = _build_supervisor_update_form(request, PrincipalRecordForm, record)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.supervisor = supervisor
        obj.save()
        messages.success(request, "تم تحديث سجل المدير/المديرة بنجاح.")
        return redirect("staffdata:records_list")

    return render(
        request,
        "staffdata/principal_form.html",
        {
            "form": form,
            "page_title": "تعديل سجل مدير/مديرة",
            "submit_label": "حفظ التعديلات",
            "is_update": True,
            "record": record,
            **_build_entry_window_context(supervisor),
        },
    )


@supervisor_login_required
def vice_update_view(request, pk):
    supervisor = get_current_supervisor(request)
    record = get_object_or_404(VicePrincipalRecord, pk=pk, supervisor=supervisor, is_active=True)

    try:
        _ensure_supervisor_can_edit(supervisor)
    except PermissionDenied as exc:
        messages.error(request, str(exc))
        return redirect("staffdata:records_list")

    form = _build_supervisor_update_form(request, VicePrincipalRecordForm, record)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.supervisor = supervisor
        obj.save()
        messages.success(request, "تم تحديث سجل الوكيل/الوكيلة بنجاح.")
        return redirect("staffdata:records_list")

    return render(
        request,
        "staffdata/vice_form.html",
        {
            "form": form,
            "page_title": "تعديل سجل وكيل/وكيلة",
            "submit_label": "حفظ التعديلات",
            "is_update": True,
            "record": record,
            **_build_entry_window_context(supervisor),
        },
    )


@supervisor_login_required
def principal_delete_view(request, pk):
    supervisor = get_current_supervisor(request)
    record = get_object_or_404(PrincipalRecord, pk=pk, supervisor=supervisor, is_active=True)

    if request.method != "POST":
        messages.info(request, "حذف السجل يتطلب طلبًا مؤكدًا.")
        return redirect("staffdata:records_list")

    try:
        _ensure_supervisor_can_delete(supervisor)
    except PermissionDenied as exc:
        messages.error(request, str(exc))
        return redirect("staffdata:records_list")

    record.is_active = False
    record.save()
    messages.success(request, "تم إلغاء تنشيط سجل المدير/المديرة بنجاح.")
    return redirect("staffdata:records_list")


@supervisor_login_required
def vice_delete_view(request, pk):
    supervisor = get_current_supervisor(request)
    record = get_object_or_404(VicePrincipalRecord, pk=pk, supervisor=supervisor, is_active=True)

    if request.method != "POST":
        messages.info(request, "حذف السجل يتطلب طلبًا مؤكدًا.")
        return redirect("staffdata:records_list")

    try:
        _ensure_supervisor_can_delete(supervisor)
    except PermissionDenied as exc:
        messages.error(request, str(exc))
        return redirect("staffdata:records_list")

    record.is_active = False
    record.save()
    messages.success(request, "تم إلغاء تنشيط سجل الوكيل/الوكيلة بنجاح.")
    return redirect("staffdata:records_list")


@supervisor_login_required
def principal_correction_request_view(request, pk):
    supervisor = get_current_supervisor(request)
    record = get_object_or_404(PrincipalRecord, pk=pk, supervisor=supervisor, is_active=True)
    return _create_correction_request(
        request,
        record,
        CorrectionRequest.TARGET_PRINCIPAL,
        "طلب تصحيح سجل مدير/مديرة",
    )


@supervisor_login_required
def vice_correction_request_view(request, pk):
    supervisor = get_current_supervisor(request)
    record = get_object_or_404(VicePrincipalRecord, pk=pk, supervisor=supervisor, is_active=True)
    return _create_correction_request(
        request,
        record,
        CorrectionRequest.TARGET_VICE,
        "طلب تصحيح سجل وكيل/وكيلة",
    )


@supervisor_login_required
def my_correction_requests_view(request):
    supervisor = get_current_supervisor(request)
    requests_qs = (
        CorrectionRequest.objects.filter(supervisor=supervisor)
        .select_related("principal_record", "vice_record")
        .order_by("-created_at")
    )

    return render(
        request,
        "staffdata/my_correction_requests.html",
        {
            "page_title": "طلبات التصحيح",
            "correction_requests": requests_qs,
            **_build_entry_window_context(supervisor),
        },
    )


@staff_member_required(login_url="/admin/login/")
def admin_overview_view(request):
    principal_base_qs = PrincipalRecord.objects.select_related("supervisor").all()
    vice_base_qs = VicePrincipalRecord.objects.select_related("supervisor").all()

    principal_records = apply_principal_filters(request, principal_base_qs)
    vice_records = apply_vice_filters(request, vice_base_qs)

    overview = build_admin_overview_context(principal_records, vice_records) or {}
    stats = overview.get("stats", {})
    supervisor_rows = overview.get("supervisor_rows", [])

    principal_total = overview.get(
        "principal_total",
        stats.get("total_principals", principal_records.count()),
    )
    vice_total = overview.get(
        "vice_total",
        stats.get("total_vices", vice_records.count()),
    )
    total_records = overview.get(
        "total_records",
        stats.get("total_records", principal_total + vice_total),
    )
    active_records = overview.get(
        "active_records",
        stats.get(
            "active_records",
            _count_active(principal_records) + _count_active(vice_records),
        ),
    )
    inactive_records = overview.get(
        "inactive_records",
        stats.get("inactive_records", total_records - active_records),
    )
    attachments_total = overview.get(
        "attachments_total",
        stats.get("attached_records", _count_attachments(principal_records)),
    )
    missing_attachments = overview.get(
        "missing_attachments",
        stats.get("missing_attachments", max(principal_total - attachments_total, 0)),
    )
    schools_total = overview.get(
        "schools_total",
        stats.get("schools_total", _collect_school_count(principal_records, vice_records)),
    )

    pending_corrections_count = CorrectionRequest.objects.filter(
        status=CorrectionRequest.STATUS_PENDING
    ).count()
    pending_account_reset_requests_count = AccountResetRequest.objects.filter(
        status=AccountResetRequest.STATUS_PENDING
    ).count()

    supervisors_count = stats.get("supervisors_count", len(supervisor_rows) or 0)

    latest_correction_requests = (
        CorrectionRequest.objects.select_related(
            "supervisor",
            "principal_record",
            "vice_record",
            "reviewed_by",
        )
        .order_by("-created_at")[:8]
    )

    context = {
        "page_title": "لوحة الإدارة العامة",
        "stats": {
            "total_records": total_records,
            "total_principals": principal_total,
            "total_vices": vice_total,
            "active_records": active_records,
            "inactive_records": inactive_records,
            "attached_records": attachments_total,
            "missing_attachments": missing_attachments,
            "schools_total": schools_total,
            "supervisors_count": supervisors_count,
            "pending_corrections_count": pending_corrections_count,
            "pending_account_reset_requests_count": pending_account_reset_requests_count,
        },
        "supervisor_rows": supervisor_rows[:10],
        "sector_rows": overview.get("sector_rows", [])[:8],
        "stage_rows": overview.get("stage_rows", [])[:8],
        "recent_records": overview.get("recent_records", [])[:10],
        "latest_correction_requests": latest_correction_requests,
        **_build_admin_filter_context(request),
        **_build_entry_window_context(),
    }

    return render(request, "staffdata/admin_overview.html", context)


@staff_member_required(login_url="/admin/login/")
def admin_supervisors_list_view(request):
    q = request.GET.get("q", "").strip()
    sector = request.GET.get("sector", "").strip()
    sort = request.GET.get("sort", "priority").strip()

    supervisors_qs = Supervisor.objects.all().order_by("full_name")

    if q:
        search_q = Q()

        if _model_has_field(Supervisor, "full_name"):
            search_q |= Q(full_name__icontains=q)
        if _model_has_field(Supervisor, "national_id"):
            search_q |= Q(national_id__icontains=q)
        if _model_has_field(Supervisor, "mobile"):
            search_q |= Q(mobile__icontains=q)
        if _model_has_field(Supervisor, "email"):
            search_q |= Q(email__icontains=q)

        supervisors_qs = supervisors_qs.filter(search_q)

    if sector and _model_has_field(Supervisor, "sector"):
        supervisors_qs = supervisors_qs.filter(sector=sector)

    supervisors = list(supervisors_qs)
    supervisor_ids = [item.id for item in supervisors]

    principals = list(
        PrincipalRecord.objects.filter(supervisor_id__in=supervisor_ids, is_active=True)
        .select_related("supervisor")
        .only("id", "supervisor_id", "created_at", "performance_file")
    )

    vices = list(
        VicePrincipalRecord.objects.filter(supervisor_id__in=supervisor_ids, is_active=True)
        .select_related("supervisor")
        .only("id", "supervisor_id", "created_at")
    )

    pending_corrections = list(
        CorrectionRequest.objects.filter(
            supervisor_id__in=supervisor_ids,
            status=CorrectionRequest.STATUS_PENDING,
        )
        .select_related("supervisor")
        .only("id", "supervisor_id", "created_at", "status")
    )

    pending_reset_requests = list(
        AccountResetRequest.objects.filter(
            supervisor_id__in=supervisor_ids,
            status=AccountResetRequest.STATUS_PENDING,
        )
        .select_related("supervisor")
        .only("id", "supervisor_id", "created_at", "status")
    )

    rows_map = defaultdict(
        lambda: {
            "principals_count": 0,
            "vices_count": 0,
            "total_count": 0,
            "attachments_count": 0,
            "missing_attachments": 0,
            "pending_corrections_count": 0,
            "pending_reset_requests_count": 0,
            "last_activity": None,
        }
    )

    for item in principals:
        row = rows_map[item.supervisor_id]
        row["principals_count"] += 1
        row["total_count"] += 1
        if getattr(item, "performance_file", None):
            row["attachments_count"] += 1
        else:
            row["missing_attachments"] += 1
        row["last_activity"] = _merge_latest_activity(row["last_activity"], item.created_at)

    for item in vices:
        row = rows_map[item.supervisor_id]
        row["vices_count"] += 1
        row["total_count"] += 1
        row["last_activity"] = _merge_latest_activity(row["last_activity"], item.created_at)

    for item in pending_corrections:
        row = rows_map[item.supervisor_id]
        row["pending_corrections_count"] += 1
        row["last_activity"] = _merge_latest_activity(row["last_activity"], item.created_at)

    for item in pending_reset_requests:
        row = rows_map[item.supervisor_id]
        row["pending_reset_requests_count"] += 1
        row["last_activity"] = _merge_latest_activity(row["last_activity"], item.created_at)

    rows = []
    for supervisor in supervisors:
        stats = rows_map.get(supervisor.id, {})
        status_label, status_class = _build_supervisor_status(
            stats.get("total_count", 0),
            stats.get("missing_attachments", 0),
            stats.get("pending_corrections_count", 0),
            stats.get("pending_reset_requests_count", 0),
        )

        rows.append(
            {
                "id": supervisor.id,
                "full_name": _safe_supervisor_value(supervisor, "full_name"),
                "national_id": _safe_supervisor_value(supervisor, "national_id"),
                "mobile": _safe_supervisor_value(supervisor, "mobile"),
                "email": _safe_supervisor_value(supervisor, "email"),
                "sector": _safe_supervisor_value(supervisor, "sector"),
                "is_active": getattr(supervisor, "is_active", True),
                "can_add_records": getattr(supervisor, "can_add_records", True),
                "can_edit_records": getattr(supervisor, "can_edit_records", False),
                "can_delete_records": getattr(supervisor, "can_delete_records", False),
                "principals_count": stats.get("principals_count", 0),
                "vices_count": stats.get("vices_count", 0),
                "total_count": stats.get("total_count", 0),
                "attachments_count": stats.get("attachments_count", 0),
                "missing_attachments": stats.get("missing_attachments", 0),
                "pending_corrections_count": stats.get("pending_corrections_count", 0),
                "pending_reset_requests_count": stats.get("pending_reset_requests_count", 0),
                "last_activity": stats.get("last_activity"),
                "status_label": status_label,
                "status_class": status_class,
            }
        )

    fallback_dt = _fallback_aware_datetime()

    if sort == "latest":
        rows.sort(
            key=lambda row: (
                row["last_activity"] or fallback_dt,
                row["total_count"],
            ),
            reverse=True,
        )
    elif sort == "activity":
        rows.sort(
            key=lambda row: (
                row["total_count"],
                row["principals_count"],
                row["vices_count"],
                row["last_activity"] or fallback_dt,
            ),
            reverse=True,
        )
    elif sort == "corrections":
        rows.sort(
            key=lambda row: (
                row["pending_corrections_count"],
                row["pending_reset_requests_count"],
                row["missing_attachments"],
                row["total_count"],
                row["last_activity"] or fallback_dt,
            ),
            reverse=True,
        )
    else:
        rows.sort(
            key=lambda row: (
                row["pending_reset_requests_count"],
                row["missing_attachments"],
                row["pending_corrections_count"],
                row["total_count"],
                row["last_activity"] or fallback_dt,
            ),
            reverse=True,
        )

    paginator = Paginator(rows, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    sector_choices = []
    if _model_has_field(Supervisor, "sector"):
        sector_choices = [
            value
            for value in Supervisor.objects.exclude(sector="")
            .exclude(sector__isnull=True)
            .values_list("sector", flat=True)
            .distinct()
            .order_by("sector")
        ]

    context = {
        "page_title": "استعراض المشرفين ومدخلاتهم",
        "page_obj": page_obj,
        "supervisors_rows": page_obj.object_list,
        "total_count": len(rows),
        "current_q": q,
        "current_sector": sector,
        "current_sort": sort,
        "sector_choices": sector_choices,
        **_build_entry_window_context(),
    }
    return render(request, "staffdata/admin_supervisors_list.html", context)


@staff_member_required(login_url="/admin/login/")
def admin_supervisor_detail_view(request, supervisor_id):
    supervisor = get_object_or_404(Supervisor, pk=supervisor_id)

    principal_records = (
        PrincipalRecord.objects.filter(supervisor=supervisor)
        .select_related("supervisor")
        .order_by("-created_at")
    )
    vice_records = (
        VicePrincipalRecord.objects.filter(supervisor=supervisor)
        .select_related("supervisor")
        .order_by("-created_at")
    )
    correction_requests = (
        CorrectionRequest.objects.filter(supervisor=supervisor)
        .select_related("principal_record", "vice_record", "reviewed_by")
        .order_by("-created_at")
    )
    account_reset_requests = (
        AccountResetRequest.objects.filter(supervisor=supervisor)
        .select_related("processed_by")
        .order_by("-created_at")
    )

    principal_count = principal_records.filter(is_active=True).count()
    vice_count = vice_records.filter(is_active=True).count()
    total_count = principal_count + vice_count
    attachments_count = _count_attachments(principal_records.filter(is_active=True))
    missing_attachments = max(principal_count - attachments_count, 0)
    pending_corrections_count = correction_requests.filter(
        status=CorrectionRequest.STATUS_PENDING
    ).count()
    pending_reset_requests_count = account_reset_requests.filter(
        status=AccountResetRequest.STATUS_PENDING
    ).count()

    last_activity = _merge_latest_activity(
        principal_records.first().created_at if principal_records.exists() else None,
        vice_records.first().created_at if vice_records.exists() else None,
        correction_requests.first().created_at if correction_requests.exists() else None,
        account_reset_requests.first().created_at if account_reset_requests.exists() else None,
    )

    status_label, status_class = _build_supervisor_status(
        total_count,
        missing_attachments,
        pending_corrections_count,
        pending_reset_requests_count,
    )

    missing_principal_records = principal_records.filter(
        is_active=True,
    ).filter(
        Q(performance_file="") | Q(performance_file__isnull=True)
    )

    recent_activity = []

    for item in principal_records[:5]:
        recent_activity.append(
            {
                "type": "مدير/مديرة",
                "title": item.full_name,
                "subtitle": item.school_name,
                "meta": item.role,
                "created_at": item.created_at,
            }
        )

    for item in vice_records[:5]:
        recent_activity.append(
            {
                "type": "وكيل/وكيلة",
                "title": item.full_name,
                "subtitle": item.school_name,
                "meta": item.role,
                "created_at": item.created_at,
            }
        )

    for item in correction_requests[:5]:
        recent_activity.append(
            {
                "type": "طلب تصحيح",
                "title": item.requested_full_name or "طلب تحديث سجل",
                "subtitle": item.get_target_type_display(),
                "meta": item.get_status_display(),
                "created_at": item.created_at,
            }
        )

    for item in account_reset_requests[:5]:
        recent_activity.append(
            {
                "type": "إعادة تهيئة",
                "title": "طلب إعادة تهيئة حساب",
                "subtitle": supervisor.full_name,
                "meta": item.get_status_display(),
                "created_at": item.created_at,
            }
        )

    recent_activity.sort(key=lambda x: x["created_at"], reverse=True)
    recent_activity = recent_activity[:12]

    context = {
        "page_title": "بطاقة المشرف",
        "supervisor": supervisor,
        "supervisor_info": {
            "full_name": _safe_supervisor_value(supervisor, "full_name"),
            "national_id": _safe_supervisor_value(supervisor, "national_id"),
            "mobile": _safe_supervisor_value(supervisor, "mobile"),
            "email": _safe_supervisor_value(supervisor, "email"),
            "sector": _safe_supervisor_value(supervisor, "sector"),
            "is_active": getattr(supervisor, "is_active", True),
            "can_add_records": getattr(supervisor, "can_add_records", True),
            "can_edit_records": getattr(supervisor, "can_edit_records", False),
            "can_delete_records": getattr(supervisor, "can_delete_records", False),
            "status_label": status_label,
            "status_class": status_class,
            "last_activity": last_activity,
        },
        "stats": {
            "principal_count": principal_count,
            "vice_count": vice_count,
            "total_count": total_count,
            "attachments_count": attachments_count,
            "missing_attachments": missing_attachments,
            "pending_corrections_count": pending_corrections_count,
            "pending_reset_requests_count": pending_reset_requests_count,
        },
        "principal_records": principal_records[:10],
        "vice_records": vice_records[:10],
        "missing_principal_records": missing_principal_records[:10],
        "correction_requests": correction_requests[:10],
        "account_reset_requests": account_reset_requests[:10],
        "pending_account_reset_request": account_reset_requests.filter(
            status=AccountResetRequest.STATUS_PENDING
        ).first(),
        "recent_activity": recent_activity,
        **_build_entry_window_context(supervisor),
    }
    return render(request, "staffdata/admin_supervisor_detail.html", context)


@staff_member_required(login_url="/admin/login/")
@transaction.atomic
def admin_supervisor_update_view(request, supervisor_id):
    supervisor = get_object_or_404(Supervisor, pk=supervisor_id)

    editable_fields = [
        field_name
        for field_name in [
            "full_name",
            "mobile",
            "email",
            "is_active",
            "can_add_records",
            "can_edit_records",
            "can_delete_records",
        ]
        if _model_has_field(Supervisor, field_name)
    ]

    form = SupervisorAdminUpdateForm(request.POST or None, instance=supervisor)

    if request.method == "POST" and form.is_valid():
        updated_supervisor = form.save(commit=False)
        update_fields = [field for field in editable_fields if field in form.changed_data]

        if update_fields:
            updated_supervisor.save(update_fields=update_fields)
            messages.success(request, "تم تحديث بيانات المشرف بنجاح.")
        else:
            messages.info(request, "لم يتم إجراء أي تغيير على بيانات المشرف.")

        return redirect("staffdata:admin_supervisor_detail", supervisor_id=supervisor.id)

    return render(
        request,
        "staffdata/admin_supervisor_form.html",
        {
            "page_title": "تعديل بيانات المشرف",
            "form": form,
            "supervisor": supervisor,
            "submit_label": "حفظ التعديلات",
            **_build_entry_window_context(),
        },
    )


@staff_member_required(login_url="/admin/login/")
def admin_principals_list_view(request):
    base_qs = PrincipalRecord.objects.select_related("supervisor").all()
    records_qs = apply_principal_filters(request, base_qs)

    paginator = Paginator(records_qs, 30)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_title": "سجلات المديرين/المديرات",
        "record_type": "principal",
        "page_obj": page_obj,
        "records": page_obj.object_list,
        "total_count": records_qs.count(),
        "export_url_name": "staffdata:admin_export_principals_csv",
        **_build_admin_filter_context(request),
        **_build_entry_window_context(),
    }
    return render(request, "staffdata/admin_records_list.html", context)


@staff_member_required(login_url="/admin/login/")
def admin_vices_list_view(request):
    base_qs = VicePrincipalRecord.objects.select_related("supervisor").all()
    records_qs = apply_vice_filters(request, base_qs)

    paginator = Paginator(records_qs, 30)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_title": "سجلات الوكلاء/الوكيلات",
        "record_type": "vice",
        "page_obj": page_obj,
        "records": page_obj.object_list,
        "total_count": records_qs.count(),
        "export_url_name": "staffdata:admin_export_vice_csv",
        **_build_admin_filter_context(request),
        **_build_entry_window_context(),
    }
    return render(request, "staffdata/admin_records_list.html", context)


@staff_member_required(login_url="/admin/login/")
def admin_correction_requests_view(request):
    qs = (
        CorrectionRequest.objects.select_related(
            "supervisor",
            "principal_record",
            "vice_record",
            "reviewed_by",
        )
        .order_by("-created_at")
    )

    status = request.GET.get("status", "").strip()
    target_type = request.GET.get("target_type", "").strip()
    q = request.GET.get("q", "").strip()

    if status:
        qs = qs.filter(status=status)

    if target_type:
        qs = qs.filter(target_type=target_type)

    if q:
        qs = qs.filter(
            Q(requested_full_name__icontains=q)
            | Q(requested_national_id__icontains=q)
            | Q(requested_mobile__icontains=q)
        )

    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "staffdata/admin_correction_requests.html",
        {
            "page_title": "طلبات التصحيح",
            "page_obj": page_obj,
            "requests": page_obj.object_list,
            "current_status": status,
            "current_target_type": target_type,
            "current_q": q,
            "status_choices": getattr(CorrectionRequest, "STATUS_CHOICES", []),
            "target_type_choices": getattr(CorrectionRequest, "TARGET_TYPE_CHOICES", []),
            **_build_entry_window_context(),
        },
    )


@staff_member_required(login_url="/admin/login/")
@transaction.atomic
def admin_correction_request_review_view(request, pk):
    correction_request = get_object_or_404(
        CorrectionRequest.objects.select_related(
            "supervisor",
            "principal_record",
            "vice_record",
        ),
        pk=pk,
    )

    if correction_request.status != CorrectionRequest.STATUS_PENDING:
        messages.warning(request, "تمت معالجة هذا الطلب مسبقًا.")
        return redirect("staffdata:admin_correction_requests")

    form = CorrectionDecisionForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        action = form.cleaned_data["action"]
        admin_note = form.cleaned_data.get("admin_note", "")

        correction_request.admin_note = admin_note
        correction_request.reviewed_by = request.user
        correction_request.reviewed_at = timezone.now()

        if action == "approve":
            _apply_correction_to_record(correction_request)
            correction_request.status = CorrectionRequest.STATUS_APPROVED
            correction_request.save()
            messages.success(request, "تم اعتماد طلب التصحيح وتحديث السجل.")
        elif action == "return":
            correction_request.status = CorrectionRequest.STATUS_RETURNED
            correction_request.save()
            messages.success(request, "تمت إعادة الطلب للمشرف للمراجعة.")
        else:
            correction_request.status = CorrectionRequest.STATUS_REJECTED
            correction_request.save()
            messages.success(request, "تم رفض طلب التصحيح.")

        return redirect("staffdata:admin_correction_requests")

    return render(
        request,
        "staffdata/admin_correction_request_review.html",
        {
            "page_title": "مراجعة طلب تصحيح",
            "correction_request": correction_request,
            "form": form,
            **_build_entry_window_context(),
        },
    )


@staff_member_required(login_url="/admin/login/")
def admin_export_principals_csv(request):
    records = apply_principal_filters(
        request,
        PrincipalRecord.objects.select_related("supervisor").all(),
    )

    headers = [
        "الاسم",
        "السجل المدني",
        "رقم الجوال",
        "الصفة",
        "المدرسة",
        "القطاع",
        "المرحلة",
        "نوع المدرسة",
        "المشرف",
        "الحالة",
        "اسم ملف الاستمارة",
        "تاريخ الإضافة",
    ]

    rows = []
    for item in records:
        rows.append(
            [
                item.full_name,
                item.national_id,
                getattr(item, "mobile", ""),
                item.role,
                item.school_name,
                item.sector,
                item.stage,
                item.school_gender,
                item.supervisor.full_name if item.supervisor else "",
                "نشط" if getattr(item, "is_active", False) else "غير نشط",
                Path(item.performance_file.name).name if item.performance_file else "",
                item.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    return _build_xlsx_response(
        filename="principals_records.xlsx",
        sheet_title="المديرون",
        headers=headers,
        rows=rows,
    )


@staff_member_required(login_url="/admin/login/")
def admin_export_vice_csv(request):
    records = apply_vice_filters(
        request,
        VicePrincipalRecord.objects.select_related("supervisor").all(),
    )

    headers = [
        "الاسم",
        "السجل المدني",
        "رقم الجوال",
        "الصفة",
        "المدرسة",
        "القطاع",
        "المرحلة",
        "نوع المدرسة",
        "المشرف",
        "الحالة",
        "تاريخ الإضافة",
    ]

    rows = []
    for item in records:
        rows.append(
            [
                item.full_name,
                item.national_id,
                getattr(item, "mobile", ""),
                item.role,
                item.school_name,
                item.sector,
                item.stage,
                item.school_gender,
                item.supervisor.full_name if item.supervisor else "",
                "نشط" if getattr(item, "is_active", False) else "غير نشط",
                item.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    return _build_xlsx_response(
        filename="vice_records.xlsx",
        sheet_title="الوكلاء",
        headers=headers,
        rows=rows,
    )


@staff_member_required(login_url="/admin/login/")
def admin_download_performance_center_view(request):
    stage_choices = [
        {"value": value, "label": label}
        for value, label in PrincipalRecord.STAGE_CHOICES
    ]

    return render(
        request,
        "staffdata/admin_download_performance_center.html",
        {
            "page_title": "تحميل الاستمارات بحسب المرحلة",
            "stage_choices": stage_choices,
            **_build_entry_window_context(),
        },
    )


@staff_member_required(login_url="/admin/login/")
def admin_download_performance_zip(request):
    records = apply_principal_filters(
        request,
        PrincipalRecord.objects.select_related("supervisor").all(),
    ).exclude(performance_file="").exclude(performance_file__isnull=True)

    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for index, item in enumerate(records, start=1):
            if not item.performance_file:
                continue

            extension = Path(item.performance_file.name).suffix or ".bin"
            school_name = _clean_filename(item.school_name)
            full_name = _clean_filename(item.full_name)
            archive_name = f"{index:03d}-{school_name}-{full_name}{extension}"

            try:
                item.performance_file.open("rb")
                zip_file.writestr(archive_name, item.performance_file.read())
            finally:
                item.performance_file.close()

    buffer.seek(0)

    stage_value = request.GET.get("stage", "").strip()
    stage_map = dict(getattr(PrincipalRecord, "STAGE_CHOICES", []))
    stage_label = stage_map.get(stage_value, "الكل")
    safe_stage = _clean_filename(stage_label)

    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = (
        f'attachment; filename="performance_forms_{safe_stage}.zip"'
    )
    return response